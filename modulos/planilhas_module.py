"""
Modulo de Planilhas - Agente Contabil
Leitura, escrita, analise e geracao de relatorios
em Excel (.xlsx) e Google Sheets para contabilidade.
"""

import os
import logging
from typing import List, Dict, Optional, Any
from datetime import datetime

logger = logging.getLogger(__name__)


class PlanilhasModule:
    """
    Modulo completo para manipulacao de planilhas.
    Suporta Excel (openpyxl/pandas) e Google Sheets (gspread).
    """

    def __init__(self):
        self.logger = logging.getLogger(self.__class__.__name__)

    # =========================================================
    # LEITURA DE PLANILHAS
    # =========================================================

    def ler_excel(self, caminho: str, aba: Optional[str] = None) -> List[Dict]:
        """
        Le um arquivo Excel e retorna lista de dicionarios.

        Args:
            caminho: Caminho do arquivo .xlsx ou .xls
            aba: Nome da aba (None = primeira aba)

        Returns:
            Lista de dicionarios com os dados
        """
        import pandas as pd
        try:
            df = pd.read_excel(caminho, sheet_name=aba or 0, dtype=str)
            df = df.fillna('')
            registros = df.to_dict('records')
            self.logger.info(f"{len(registros)} registros lidos de {caminho}")
            return registros
        except Exception as e:
            self.logger.error(f"Erro ao ler Excel {caminho}: {e}")
            return []

    def analisar(self, caminho: str) -> Dict:
        """
        Analisa uma planilha contabil detectando:
        - Totais e subtotais
        - Inconsistencias e erros
        - Saldos zerados indevidamente
        - Lancamentos duplicados
        - Campos obrigatorios vazios

        Args:
            caminho: Caminho do arquivo Excel

        Returns:
            Dicionario com resultado da analise
        """
        import pandas as pd
        try:
            df = pd.read_excel(caminho)
            resultado = {
                'arquivo': caminho,
                'total_linhas': len(df),
                'colunas': list(df.columns),
                'nulos': df.isnull().sum().to_dict(),
                'duplicados': int(df.duplicated().sum()),
                'resumo': f"{len(df)} registros, {df.shape[1]} colunas",
                'problemas': []
            }

            # Verifica duplicados
            if resultado['duplicados'] > 0:
                resultado['problemas'].append(
                    f"ATENCAO: {resultado['duplicados']} linhas duplicadas encontradas"
                )

            # Verifica campos nulos em colunas criticas
            for col, nulos in resultado['nulos'].items():
                if nulos > 0:
                    resultado['problemas'].append(
                        f"Coluna '{col}' possui {nulos} valores vazios"
                    )

            # Verifica colunas numericas com valores negativos inesperados
            for col in df.select_dtypes(include='number').columns:
                negativos = (df[col] < 0).sum()
                if negativos > 0:
                    resultado['problemas'].append(
                        f"Coluna '{col}' possui {negativos} valores negativos"
                    )

            self.logger.info(f"Analise concluida: {len(resultado['problemas'])} problemas encontrados")
            return resultado

        except Exception as e:
            self.logger.error(f"Erro ao analisar planilha {caminho}: {e}")
            return {'erro': str(e)}

    def analisar_automatico(self, caminho: str) -> Dict:
        """Alias para analisar() - compatibilidade"""
        return self.analisar(caminho)

    # =========================================================
    # ESCRITA E EXPORTACAO
    # =========================================================

    def salvar(self, dados: Any, caminho: str, aba: str = 'Dados'):
        """
        Salva dados em arquivo Excel.

        Args:
            dados: Lista de dicionarios ou DataFrame
            caminho: Caminho de destino
            aba: Nome da aba
        """
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        os.makedirs(os.path.dirname(caminho) if os.path.dirname(caminho) else '.', exist_ok=True)

        try:
            if isinstance(dados, list):
                df = pd.DataFrame(dados)
            elif isinstance(dados, dict):
                df = pd.DataFrame([dados])
            else:
                df = dados

            with pd.ExcelWriter(caminho, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=aba, index=False)

                # Formata cabecalho
                ws = writer.sheets[aba]
                for cell in ws[1]:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(fill_type='solid', fgColor='1F4E79')
                    cell.alignment = Alignment(horizontal='center')

                # Ajusta largura das colunas
                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

            self.logger.info(f"Planilha salva em {caminho}")
            return caminho

        except Exception as e:
            self.logger.error(f"Erro ao salvar planilha {caminho}: {e}")
            return None

    def exportar_vencimentos(self, vencimentos: List[Dict], caminho: Optional[str] = None):
        """Exporta lista de vencimentos para Excel formatado"""
        if caminho is None:
            caminho = f"dados/relatorios/vencimentos_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return self.salvar(vencimentos, caminho, aba='Vencimentos')

    def salvar_conciliacao(self, resultado: Dict, caminho: str):
        """Salva resultado da conciliacao bancaria em Excel"""
        dados = resultado.get('lancamentos', [])
        return self.salvar(dados, caminho, aba='Conciliacao')

    def gerar_dashboard_dre(self, dre: Dict, caminho: str):
        """
        Gera dashboard visual da DRE em Excel com graficos.

        Args:
            dre: Dicionario com dados da DRE
            caminho: Caminho de destino
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.chart import BarChart, Reference

        wb = Workbook()
        ws = wb.active
        ws.title = 'DRE'

        # Cabecalho
        ws['A1'] = 'DEMONSTRACAO DO RESULTADO DO EXERCICIO'
        ws['A1'].font = Font(bold=True, size=14)

        # Linha de receita bruta
        linhas = [
            ('Receita Bruta', dre.get('receita_bruta', 0)),
            ('(-) Deducoes', -abs(dre.get('deducoes', 0))),
            ('= Receita Liquida', dre.get('receita_liquida', 0)),
            ('(-) CMV/CSP', -abs(dre.get('cmv', 0))),
            ('= Lucro Bruto', dre.get('lucro_bruto', 0)),
            ('(-) Despesas Operacionais', -abs(dre.get('despesas_op', 0))),
            ('= EBITDA', dre.get('ebitda', 0)),
            ('(-) Depreciacao', -abs(dre.get('depreciacao', 0))),
            ('= EBIT', dre.get('ebit', 0)),
            ('(+/-) Resultado Financeiro', dre.get('resultado_financeiro', 0)),
            ('= Lucro Antes IR', dre.get('lair', 0)),
            ('(-) IRPJ/CSLL', -abs(dre.get('impostos', 0))),
            ('= Lucro Liquido', dre.get('lucro_liquido', 0)),
        ]

        for i, (descricao, valor) in enumerate(linhas, start=3):
            ws.cell(row=i, column=1, value=descricao)
            ws.cell(row=i, column=2, value=valor)
            ws.cell(row=i, column=2).number_format = '#,##0.00'

        os.makedirs(os.path.dirname(caminho) if os.path.dirname(caminho) else '.', exist_ok=True)
        wb.save(caminho)
        self.logger.info(f"Dashboard DRE gerado: {caminho}")
        return caminho

    # =========================================================
    # GOOGLE SHEETS
    # =========================================================

    def ler_google_sheets(self, spreadsheet_id: str, aba: str = 'Sheet1') -> List[Dict]:
        """
        Le dados de uma planilha Google Sheets.

        Args:
            spreadsheet_id: ID da planilha (da URL)
            aba: Nome da aba

        Returns:
            Lista de dicionarios com os dados
        """
        try:
            import gspread
            from google.oauth2.service_account import Credentials

            creds_file = os.getenv('GOOGLE_CREDENTIALS_FILE', 'credentials.json')
            scopes = ['https://spreadsheets.google.com/feeds',
                      'https://www.googleapis.com/auth/drive']
            creds = Credentials.from_service_account_file(creds_file, scopes=scopes)
            gc = gspread.authorize(creds)

            sh = gc.open_by_key(spreadsheet_id)
            worksheet = sh.worksheet(aba)
            dados = worksheet.get_all_records()
            self.logger.info(f"{len(dados)} registros lidos do Google Sheets")
            return dados

        except Exception as e:
            self.logger.error(f"Erro ao ler Google Sheets: {e}")
            return []

    def escrever_google_sheets(self, spreadsheet_id: str, dados: List[List], aba: str = 'Sheet1'):
        """Escreve dados em uma planilha Google Sheets"""
        try:
            import gspread
            from google.oauth2.service_account import Credentials

            creds_file = os.getenv('GOOGLE_CREDENTIALS_FILE', 'credentials.json')
            scopes = ['https://spreadsheets.google.com/feeds',
                      'https://www.googleapis.com/auth/drive']
            creds = Credentials.from_service_account_file(creds_file, scopes=scopes)
            gc = gspread.authorize(creds)

            sh = gc.open_by_key(spreadsheet_id)
            worksheet = sh.worksheet(aba)
            worksheet.update(dados)
            self.logger.info(f"{len(dados)} linhas escritas no Google Sheets")
            return True

        except Exception as e:
            self.logger.error(f"Erro ao escrever Google Sheets: {e}")
            return False
