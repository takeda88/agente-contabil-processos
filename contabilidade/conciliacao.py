"""Modulo de conciliacao bancaria.

TODO: REFATORAR - Substituir difflib por logica financeira adequada.
      A reconciliacao deve usar:
      - Correspondencia exata de valores (Decimal)
      - Janela de datas (+/- 3 dias)
      - Matching por historico/descricao
      Ver RELATORIO_MELHORIAS.md secao 4.1 para detalhes.
"""

"""
Modulo de Conciliacao Bancaria.
Cruza extratos bancarios com lancamentos contabeis e identifica divergencias.
"""

import os
import logging
from typing import Dict, List
from difflib import SequenceMatcher
from openpyxl import Workbook


class ConciliacaoBancaria:
    """
    Modulo responsavel por conciliacao bancaria.
    """

    def __init__(self, classificador=None):
        self.classificador = classificador
        self.logger = logging.getLogger(self.__class__.__name__)
        self.logger.info("ConciliacaoBancaria inicializado")

    def conciliar(self, extrato_bancario: List[Dict], lancamentos_contabeis: List[Dict]) -> Dict:
        """
        Cruza extrato bancario com lancamentos contabeis.

        Args:
            extrato_bancario: Lista de lancamentos do extrato
            lancamentos_contabeis: Lista de lancamentos contabeis

        Returns:
            Dicionario com resultado da conciliacao
        """
        try:
            self.logger.info(f"Conciliando: {len(extrato_bancario)} extrato x {len(lancamentos_contabeis)} contabil")
            
            conciliados = []
            nao_conciliados_banco = []
            nao_conciliados_contabil = list(lancamentos_contabeis)
            
            for lanc_banco in extrato_bancario:
                melhor_match = None
                maior_similaridade = 0
                
                for i, lanc_contabil in enumerate(nao_conciliados_contabil):
                    similaridade = self._comparar_lancamentos(lanc_banco, lanc_contabil)
                    
                    if similaridade > maior_similaridade and similaridade >= 0.7:
                        maior_similaridade = similaridade
                        melhor_match = (i, lanc_contabil)
                
                if melhor_match:
                    conciliados.append({
                        'banco': lanc_banco,
                        'contabil': melhor_match[1],
                        'similaridade': maior_similaridade
                    })
                    nao_conciliados_contabil.pop(melhor_match[0])
                else:
                    nao_conciliados_banco.append(lanc_banco)
            
            resultado = {
                'total_extrato': len(extrato_bancario),
                'total_contabil': len(lancamentos_contabeis),
                'conciliados': len(conciliados),
                'divergentes_banco': len(nao_conciliados_banco),
                'divergentes_contabil': len(nao_conciliados_contabil),
                'detalhes_conciliados': conciliados,
                'detalhes_divergentes_banco': nao_conciliados_banco,
                'detalhes_divergentes_contabil': nao_conciliados_contabil
            }
            
            self.logger.info(f"Conciliacao concluida: {len(conciliados)} conciliados")
            return resultado

        except Exception as e:
            self.logger.error(f"Erro na conciliacao: {e}")
            return {'erro': str(e)}

    def identificar_divergencias(self, resultado: Dict) -> List[Dict]:
        """
        Lista itens nao conciliados.

        Args:
            resultado: Resultado da conciliacao

        Returns:
            Lista de divergencias
        """
        try:
            divergencias = []
            
            for item in resultado.get('detalhes_divergentes_banco', []):
                divergencias.append({
                    'tipo': 'NAO_ENCONTRADO_CONTABILIDADE',
                    'origem': 'BANCO',
                    'data': item.get('data'),
                    'descricao': item.get('descricao'),
                    'valor': item.get('valor')
                })
            
            for item in resultado.get('detalhes_divergentes_contabil', []):
                divergencias.append({
                    'tipo': 'NAO_ENCONTRADO_BANCO',
                    'origem': 'CONTABILIDADE',
                    'data': item.get('data'),
                    'descricao': item.get('descricao'),
                    'valor': item.get('valor')
                })
            
            self.logger.info(f"{len(divergencias)} divergencias identificadas")
            return divergencias

        except Exception as e:
            self.logger.error(f"Erro ao identificar divergencias: {e}")
            return []

    def calcular_saldo(self, lancamentos: List[Dict]) -> float:
        """
        Calcula saldo (creditos - debitos).

        Args:
            lancamentos: Lista de lancamentos

        Returns:
            Saldo calculado
        """
        try:
            saldo = 0.0
            
            for lanc in lancamentos:
                valor = self._normalizar_valor(lanc.get('valor', 0))
                tipo = lanc.get('tipo', '').upper()
                
                if tipo == 'CREDITO' or tipo == 'C':
                    saldo += valor
                elif tipo == 'DEBITO' or tipo == 'D':
                    saldo -= valor
                else:
                    # Se nao tem tipo, usa o sinal do valor
                    saldo += valor
            
            self.logger.info(f"Saldo calculado: R$ {saldo:.2f}")
            return round(saldo, 2)

        except Exception as e:
            self.logger.error(f"Erro ao calcular saldo: {e}")
            return 0.0

    def exportar_relatorio(self, resultado: Dict, destino: str) -> bool:
        """
        Exporta conciliacao para Excel.

        Args:
            resultado: Resultado da conciliacao
            destino: Caminho do arquivo Excel

        Returns:
            True se sucesso
        """
        try:
            self.logger.info(f"Exportando relatorio: {destino}")
            
            wb = Workbook()
            wb.remove(wb.active)
            
            # Aba Resumo
            ws_resumo = wb.create_sheet("Resumo")
            ws_resumo.append(["Total Extrato", resultado['total_extrato']])
            ws_resumo.append(["Total Contabil", resultado['total_contabil']])
            ws_resumo.append(["Conciliados", resultado['conciliados']])
            ws_resumo.append(["Divergentes Banco", resultado['divergentes_banco']])
            ws_resumo.append(["Divergentes Contabil", resultado['divergentes_contabil']])
            
            # Aba Conciliados
            ws_conc = wb.create_sheet("Conciliados")
            ws_conc.append(["Data Banco", "Desc Banco", "Valor Banco", "Data Contabil", "Desc Contabil", "Valor Contabil", "Similaridade"])
            
            for item in resultado['detalhes_conciliados']:
                banco = item['banco']
                contabil = item['contabil']
                ws_conc.append([
                    banco.get('data'),
                    banco.get('descricao'),
                    banco.get('valor'),
                    contabil.get('data'),
                    contabil.get('descricao'),
                    contabil.get('valor'),
                    f"{item['similaridade']:.0%}"
                ])
            
            # Aba Divergencias
            ws_div = wb.create_sheet("Divergencias")
            ws_div.append(["Origem", "Data", "Descricao", "Valor"])
            
            for item in resultado['detalhes_divergentes_banco']:
                ws_div.append(["BANCO", item.get('data'), item.get('descricao'), item.get('valor')])
            
            for item in resultado['detalhes_divergentes_contabil']:
                ws_div.append(["CONTABIL", item.get('data'), item.get('descricao'), item.get('valor')])
            
            os.makedirs(os.path.dirname(destino), exist_ok=True)
            wb.save(destino)
            
            self.logger.info(f"Relatorio exportado: {destino}")
            return True

        except Exception as e:
            self.logger.error(f"Erro ao exportar relatorio: {e}")
            return False

    def _normalizar_valor(self, valor) -> float:
        """
        Converte string de valor para float.

        Args:
            valor: Valor em qualquer formato

        Returns:
            Valor como float
        """
        if isinstance(valor, (int, float)):
            return float(valor)
        
        if isinstance(valor, str):
            # Remove R$, espacos e converte , em .
            valor = valor.replace('R$', '').replace(' ', '')
            valor = valor.replace('.', '').replace(',', '.')
            
            try:
                return float(valor)
            except ValueError:
                return 0.0
        
        return 0.0

    def _comparar_lancamentos(self, banco: Dict, contabil: Dict) -> float:
        """
        Retorna % de similaridade entre dois lancamentos.

        Args:
            banco: Lancamento do banco
            contabil: Lancamento contabil

        Returns:
            Similaridade (0 a 1)
        """
        score = 0.0
        
        # Compara valores (peso 50%)
        valor_banco = self._normalizar_valor(banco.get('valor', 0))
        valor_contabil = self._normalizar_valor(contabil.get('valor', 0))
        
        if abs(valor_banco - valor_contabil) <= self.margem_tolerancia:
            score += 0.5
        
        # Compara datas (peso 30%)
        if banco.get('data') == contabil.get('data'):
            score += 0.3
        
        # Compara descricoes (peso 20%)
        desc_banco = str(banco.get('descricao', '')).lower()
        desc_contabil = str(contabil.get('descricao', '')).lower()
        
        similaridade_desc = SequenceMatcher(None, desc_banco, desc_contabil).ratio()
        score += 0.2 * similaridade_desc
        
        return score
